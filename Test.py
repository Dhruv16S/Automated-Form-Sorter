import openpyxl
a = "Hello, <<Name>>. Thank you for choosing the following options - <<Choices>>"
wb = openpyxl.load_workbook("Sample.xlsx")
mailbody_check = {}
sheet = wb.active
for i in range(1, sheet.max_column + 1):
    if a.find(f"<<{sheet.cell(1, i).value}>>") != -1:
        mailbody_check[f"<<{sheet.cell(1, i).value}>>"] = i
s = ""
for i in range(2, sheet.max_row + 1):
    a = "Hello, <<Name>>. Thank you for choosing the following options - <<Choices>>"
    for field_name, column_number in mailbody_check.items():
        a = a.replace(field_name, sheet.cell(row = i, column = column_number).value)
    print(a)