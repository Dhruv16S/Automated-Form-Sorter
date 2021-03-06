import tkinter
from tkinter.constants import BOTTOM, E, END, GROOVE, LEFT, NE, NW, RAISED, RIGHT, SE, TOP, Y
import openpyxl
from tkinter import filedialog
from PIL import Image, ImageTk
from selenium import webdriver
import json
import re 
import smtplib
import random
import string
from email.message import EmailMessage
from email.mime.text import MIMEText
import os
#insert your Chrome Driver path here
insert_your_path = "C:/Chrome Driver/chromedriver"
WINDOWBG1 = "#ffd9e2"
WINDOWBG2 = "#ceffff"
WINDOWBG3 = "#fff0c2"
WINDOWBG4 = "#ffda52"
regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
h = 500; w = 700
window = tkinter.Tk()
ws = window.winfo_screenwidth() 
hs = window.winfo_screenheight() 
x = (ws/2) - (w/2)
y = (hs/2) - (h/2)
window.title("Sortle")
window.minsize(height = h, width = w)
window.geometry('%dx%d+%d+%d' % (w, h, x, y))
window.config(background = WINDOWBG2)

def Remove():
    widgets = window.winfo_children()
    for item in widgets :
        if item.winfo_children() :
            widgets.extend(item.winfo_children())
    for item in widgets:
        item.destroy()

def Form_Sorting():
    Remove()
    class TkinterReturns_From:
        def __init__(self):
            self.column = 0
            self.n = 0
            self.notcreated = []
            self.finalmsg = tkinter.Label()
            self.dropdown = tkinter.Label()
            self.helpbutton = tkinter.Button()
            self.Fields_available = tkinter.Listbox()

        def Fields_Received(self):
            self.fields_fromuser = self.custom_choice.get(1.0, END)
            Creating_Files(self.column)

        def Option_Selected(self):
            if self.radio_state.get() == 1:
                self.finalmsg.destroy()
                self.dropdown.destroy()
                self.helpbutton.destroy()
                if self.custom_choice.winfo_ismapped():
                    self.custom_choice.destroy()
                    self.submit_choices.destroy()
                Creating_Files(self.column)
            elif self.radio_state.get() == 2:
                self.finalmsg.destroy()
                def RemoveHelp():
                    fields.help.destroy()
                def button_clicked():
                    fields.n += 1
                    if(fields.n % 2 != 0):
                        fields.help = tkinter.Text(height = 5, width = 34,background = WINDOWBG3, font = ("Consolas",7,"bold"), borderwidth = 0)
                        fields.help.insert(END, "' , ' separated values are \nconsidered in the same file\n' ; ' separated values are \nconsidered as different files")
                        fields.help.place(x = 0, y = 395)
                        window.mainloop()
                    elif(fields.n % 2 == 0):
                        RemoveHelp()

                fields.helpimg = Image.open("Images/Help.png")
                fields.helpimg = ImageTk.PhotoImage(fields.helpimg)
                fields.helpbutton=tkinter.Button(image = fields.helpimg,command=button_clicked, borderwidth = 0, background = WINDOWBG2)
                fields.helpbutton.place(x = 168, y = 450)

                self.custom_choice = tkinter.Text(width = 30, height = 1, font = ("Consolas",12,"bold"), borderwidth = 0)
                self.custom_choice.place(in_ = self.Fields_available, relx = 1.0, x = 5, y = 140)
                self.custom_choice.focus()
                self.submit_choices = tkinter.Button(text = "Submit", command = fields.Fields_Received)
                self.submit_choices.place(in_ = self.Fields_available, relx = 1.0, x = 286, y = 139)

        def Field_Selected(self,event):
            self.choices_available = [ ]
            self.column = self.Sheet_Parameters[self.Fields_available.get(self.Fields_available.curselection())]
            for row in range(2, self.sheet.max_row + 1):
                cell_choices = str(self.sheet.cell(row, self.column).value).split(",")
                cell_choices = [option.strip() for option in cell_choices]
                for option in cell_choices:
                    if option not in self.choices_available:
                        self.choices_available.append(option)
            try:
                self.choices_available.remove("None")
            except:
                pass
            self.scrollbar = tkinter.Scrollbar(window)
            self.choices_available_textbox = tkinter.Text(height = 5, width = 50, font = ("Consolas",10,"bold"), borderwidth = 0, yscrollcommand = self.scrollbar.set)
            self.scrollbar.place(in_ = self.choices_available_textbox, relx = 1.0)
            self.scrollbar.config(command = self.choices_available_textbox.yview)
            for options in self.choices_available : 
                self.choices_available_textbox.insert(END, u'\u2022 {}\n'.format(options))
            self.choices_available_textbox.place(in_ = self.Fields_available, relx = 1.0, rely = 0)
            self.radio_state = tkinter.IntVar()
            self.radiobutton1 = tkinter.Radiobutton(text = "Create a New Excel File for every distinct value encountered ", value = 1, variable = self.radio_state, command = fields.Option_Selected, background = WINDOWBG2)
            self.radiobutton2 = tkinter.Radiobutton(text = "Create an Excel File for fields that contain : ", value = 2, variable = self.radio_state, command = fields.Option_Selected, background = WINDOWBG2)
            self.radiobutton1.place(in_ = self.Fields_available, relx = 1.0, x = 5, y = 80)
            self.radiobutton2.place(in_ = self.Fields_available, relx = 1.0, x = 5, y = 110)
            self.custom_choice = tkinter.Text(width = 10, height = 1,font = ("Consolas",14,"bold"), borderwidth = 0)          

        def Open_Drive(self):
            chrome_driver_path = insert_your_path
            driver = webdriver.Chrome(executable_path = chrome_driver_path)
            driver.get("https://accounts.google.com/signin/v2/identifier?service=writely&sacu=1&rip=1&flowName=GlifWebSignIn&flowEntry=ServiceLogin")
            login = driver.find_element_by_id("identifierId")
            emailid = on_home.username
            login.send_keys(emailid)
            next_button = driver.find_element_by_xpath('//*[@id="identifierNext"]/div/button/span')
            next_button.click()

        def browse_file(self):
            self.Fields_available.destroy()
            setattr(fields,'name',filedialog.askopenfilename(filetypes = (("Excel Files", "*.xlsx"),("All files", "*"))))
            wb = openpyxl.load_workbook(fields.name)
            if file_name.get(1.0,END) != "":
                file_name.delete(1.0,END)
            file_name.insert(END,f"{fields.name}")
            self.sheet = wb.active 
            self.Sheet_Parameters = {}
            self.column = self.sheet.max_column
            self.max_column = self.sheet.max_column
            self.row = self.sheet.max_row
            for i in range(1, self.column + 1):
                self.Sheet_Parameters[self.sheet.cell(row = 1, column = i).value] = i

            label = tkinter.Label(text = "The following fields have been identified.\nHow would you like to classify the form : ", background = WINDOWBG2).place(x = 40, y = 242)
            self.Fields_available = tkinter.Listbox(height = len(self.Sheet_Parameters), borderwidth=0, font = ("Consolas",10,"bold"))
            for key, value in self.Sheet_Parameters.items():
                self.Fields_available.insert(value, key)
                self.Fields_available.bind("<<ListboxSelect>>",fields.Field_Selected)
                self.Fields_available.place(x = 40, y = 290)

    fields = TkinterReturns_From()

    fields.browse_img = Image.open("Images/Browse.png").resize((150,150))
    fields.browse_img = ImageTk.PhotoImage(fields.browse_img)
    fields.browse_Button = tkinter.Button(image = fields.browse_img, command = fields.browse_file, borderwidth = 0, background = WINDOWBG2)
    fields.browse_Button.place(x = 175, y = 40)
    browselabel = tkinter.Label(text = "Browse Files", background = WINDOWBG2).place(x = 200, y = 175)
    fields.drive_img = Image.open("Images/GDrive.png").resize((150,150))
    fields.drive_img = ImageTk.PhotoImage(fields.drive_img)
    fields.download_Button = tkinter.Button(image = fields.drive_img, command = fields.Open_Drive,borderwidth = 0, background = WINDOWBG2)
    fields.download_Button.place(x = 350, y = 40)
    drivelabel = tkinter.Label(text = "Download From Drive", background = WINDOWBG2).place(x = 360, y = 175)
    openfile_label = tkinter.Label(text = "File Opened : ", background = WINDOWBG2).place(x = 125, y = 202)
    file_name = tkinter.Text(height = 1, width = 80, font = ("Consolas",8,"bold"), borderwidth = 0)
    file_name.place(x = 210, y = 205)
    fields.home_img = Image.open("Images/Home.png")
    fields.home_img = ImageTk.PhotoImage(fields.home_img)
    homebutton = tkinter.Button(image = fields.home_img, command = on_home.MainMenu, background = WINDOWBG2, borderwidth = 0)
    homebutton.pack(side = TOP, anchor = NE)
    loggedin = tkinter.Label(text = f"Logged in as {on_home.username}", background = WINDOWBG2).pack(side = BOTTOM, anchor = SE)
    def Creating_Files(operating_column):
        fields.finalmsg.destroy()
        fields.dropdown.destroy()
        if fields.radio_state.get() == 1:
            for i in range(2, fields.row + 1):    
                choices_opted = str(fields.sheet.cell(i, operating_column).value).split(",")      
                file_name = [choice.strip() for choice in choices_opted]
                for field_types in file_name:
                    if field_types == "None":
                        continue                                 
                    try:
                        new_wb = openpyxl.load_workbook(f"{field_types}.xlsx")
                    except FileNotFoundError:
                        new_wb = openpyxl.Workbook()
                        new_sheet_fields = new_wb.active
                        for k in range(1, fields.max_column + 1):
                            new_sheet_fields.cell(row = 1, column = k, value = fields.sheet.cell(row = 1, column = k).value)
                    new_sheet = new_wb.active
                    new_rows = new_sheet.max_row + 1
                    for j in range(1, fields.max_column + 1):
                        if j == operating_column:
                            new_sheet.cell(new_rows,j,value = field_types)
                        else:
                            new_sheet.cell(new_rows,j,value = fields.sheet.cell(i,j).value)    
                        new_wb.save(f"{field_types}.xlsx")

        else:
            users_list = fields.fields_fromuser.split(",")
            users_criteria = [choices.strip() for choices in users_list]
            users_multiplelist = fields.fields_fromuser.split(";")
            users_multiplecriteria = [choices.strip() for choices in users_multiplelist]
            for choice_name in users_multiplecriteria:
                for i in range(2, fields.row + 1):
                    choices_opted = str(fields.sheet.cell(i, operating_column).value).split(",")      
                    file_name = [choice.strip() for choice in choices_opted]
                    choice_name_indv = choice_name.split(",")
                    if set(choice_name_indv).issubset(set(file_name)):
                        excel_file = ",".join([str(characters) for characters in choice_name_indv])
                        try:
                            new_wb = openpyxl.load_workbook(f"{excel_file}.xlsx")
                        except FileNotFoundError:
                            new_wb = openpyxl.Workbook()
                            new_sheet_fields = new_wb.active
                            for k in range(1, fields.max_column + 1):
                                new_sheet_fields.cell(row = 1, column = k, value = fields.sheet.cell(row = 1, column = k).value)
                        new_sheet = new_wb.active
                        new_rows = new_sheet.max_row + 1                          
                        for j in range(1, fields.max_column + 1):
                            if j == operating_column:
                                new_sheet.cell(new_rows,j,value = excel_file)
                            else:
                                new_sheet.cell(new_rows,j,value = fields.sheet.cell(i,j).value)    
                            new_wb.save(f"{excel_file}.xlsx")
            not_created = []
            for files in users_multiplelist:
                file = files.strip("\n")
                if not os.path.exists(f"{file}.xlsx"):
                    not_created.append(file)
            fields.__setattr__("notcreated",not_created)
        if len(fields.notcreated) == 0:
            fields.finalmsg = tkinter.Label(text = "Process Complete", font = ("Consolas"), background = WINDOWBG2)
            fields.finalmsg.place(x = 275, y = 450)
        else:
            fields.finalmsg = tkinter.Label(text = f"Following files were not created : ", font = ("Consolas"), background = WINDOWBG2)
            fields.finalmsg.place(x = 200, y = 450)
            current = tkinter.StringVar()
            current.set(fields.notcreated[0])
            fields.dropdown = tkinter.OptionMenu(window, current, *fields.notcreated)
            fields.dropdown.place(x = 560, y = 450)
    
def Send_Mails():
    Remove()
    class TkinterReturns_Mail:
        def __init__(self):
            self.state = tkinter.IntVar()
            self.file_path = "../"
            self.nextpage_img = "../"
            self.msg = EmailMessage()
            self.positions2 = {}
            self.widgets2 = []
            self.createpage2 = True
            self.next_button = tkinter.Button()
            self.visitednextpage = False
            self.visitedprevious = False
            self.attachment_path = "../"
            self.n = 0
            self.completed_label = tkinter.Label()
            self.wrong_password = tkinter.Label()
            self.notfilled = ["var"]
            self.subject = tkinter.Text()
            self.mail_body = tkinter.Text()
            self.yesorno = tkinter.IntVar()
            self.confirmation = tkinter.Label()
            self.yes_radio = tkinter.Radiobutton()
            self.no_radio = tkinter.Radiobutton()
            self.fieldsnotfilled = tkinter.StringVar()
            self.fieldsnotfilled = self.notfilled[0]
            self.field_options = tkinter.OptionMenu(window, self.fieldsnotfilled, *self.notfilled)

        def BrowseAttachments(self):
            self.attachment_path = filedialog.askopenfilename(filetypes = (("All files", "*"), ("Template files", "*.type")))
            if self.attachment_name.get(1.0, END) != "":
                self.attachment_name.delete(1.0, END) 
            self.attachment_name.insert(END,f"{self.attachment_path}")

        def Page2Options(self):
            if self.state.get():
                self.browse_attachment = tkinter.Button(text = "Browse", borderwidth = 0, command = self.BrowseAttachments, background = WINDOWBG2)
                self.browse_attachment.place(x = 55, y = 125)
                self.attachment_name = tkinter.Text(height = 1, width = 75, font = ("Consolas",10,"bold"), borderwidth = 0)
                self.attachment_name.place(x = 105, y = 127) 
            else:
                self.browse_attachment.destroy()
                self.attachment_name.destroy()           

        def NextPage(self):
            if not self.visitednextpage:
                self.positions = {}
                self.widgets = window.winfo_children()
                for item in self.widgets :
                    if item.winfo_children() :
                        self.widgets.extend(item.winfo_children())
                for item in self.widgets:
                    self.positions[item] = [item.winfo_x(), item.winfo_y()]
                self.visitednextpage = True
            for item in self.widgets:
                try :
                    item.place_forget()
                except:
                    pass
            for item in self.widgets2:
                try:
                    item.place(x = self.positions2[item][0], y = self.positions2[item][1])
                except:
                    pass
            if self.createpage2:
                loggedin = tkinter.Label(text = f"Logged in as {on_home.username}", background = WINDOWBG2).pack(side = BOTTOM, anchor = SE)
                self.state = tkinter.IntVar()
                self.add_attachments = tkinter.Checkbutton(text = "Add Attachments", variable = self.state, command = self.Page2Options, background = WINDOWBG2)
                self.add_attachments.place(x = 50, y = 100)
                #Not adding additional conditions, as Excel Sheet already classifies and it will be easier to add additional details there itself.
                self.previouspage_img = Image.open("Images/PreviousPage.png").resize((50,50))
                self.previouspage_img = ImageTk.PhotoImage(self.previouspage_img)
                self.previouspage = tkinter.Button(image = self.previouspage_img, command = self.PreviousPage, borderwidth = 0, background = WINDOWBG2)
                self.previouspage.place(x = 0, y = 250)
                self.send_mailsimg = Image.open("Images/SendMails.png")
                self.send_mailsimg = ImageTk.PhotoImage(self.send_mailsimg)
                mail_preview = tkinter.Button(image = self.send_mailsimg, borderwidth = 0, command = Write_Mails, background = WINDOWBG2)   #or mail preview
                mail_preview.place(x = 50, y = 165)
                mailpreview_label = tkinter.Label(text = "Send Mails", background = WINDOWBG2).place(x = 50, y = 249)
                
                

        def PreviousPage(self):
            if not self.visitedprevious:
                self.positions2 = {}
                self.widgets2 = window.winfo_children()
                for item in self.widgets2 :
                    if item.winfo_children() :
                        self.widgets2.extend(item.winfo_children())
                for item in self.widgets2:
                    self.positions2[item] = [item.winfo_x(), item.winfo_y()]
                if len(list(set(self.widgets2).intersection(set(self.widgets)))) == 0:
                    self.widgets2 = self.widgets
                else:
                    self.widgets2 = list(set(self.widgets2).union(set(self.widgets)) - set(self.widgets2).intersection(set(self.widgets)))
                self.visitedprevious = True
            for items in self.widgets2:
                try:
                    items.place_forget()
                except:
                    pass
            mails.confirmation.destroy()
            mails.yes_radio.destroy()
            mails.no_radio.destroy()
            mails.field_options.destroy()
            mails.wrong_password.destroy()
            mails.completed_label.destroy()
            for item in self.widgets:
                if self.positions[item][0] == self.positions[item][1] == 0:
                    continue
                try : 
                    item.place(x = self.positions[item][0], y = self.positions[item][1])
                except:
                    pass
            self.createpage2 = False

    mails = TkinterReturns_Mail()

    def DisplayGmailPassword():
        if state_check.get() == 1:
            gmail_password.config(show = "")
        else:
            gmail_password.config(show = "*")

    def Browse_File():
        if file_name.get(1.0,END) != "":
            file_name.delete(1.0,END)
        setattr(mails, "file_path", filedialog.askopenfilename(filetypes = ( ("Excel files", "*.xlsx"), ("All files", "*"))))
        file_name.insert(END, f"{mails.file_path}")
        workbook = openpyxl.load_workbook(mails.file_path)
        sheet = workbook.active
        col = sheet.max_column
        field_list = []
        for i in range(1, col + 1):
            field_list.append(sheet.cell(row = 1, column = i).value)
        setattr(mails, "fields_available", field_list)
        if displaying_fields.get(1.0, END) != "":
            displaying_fields.delete(1.0, END)
        for name in mails.fields_available:
            displaying_fields.insert(END, u'\u2022 {}\n'.format(name))

    def Write_Mails_Final():
        wb = openpyxl.load_workbook(mails.file_path)
        sheet = wb.active
        email_message = mails.mail_body.get(1.0, END)
        addresses = to_address.get(1.0, END)
        mailbody_check = {}
        for i in range(1, sheet.max_column + 1):
            if email_message.find(f"<<{sheet.cell(1, i).value.strip()}>>") != -1:
                mailbody_check[f"<<{sheet.cell(1, i).value.strip()}>>"] = i
        for i in range(1, sheet.max_column + 1):
            if addresses.find(f"<<{sheet.cell(1, i).value.strip()}>>") != -1:
                mailbody_check["emailid"] = i
        with smtplib.SMTP("smtp.gmail.com") as connection:
            connection.starttls()
            try:
                connection.login(user = from_address.get(1.0, END).strip(), password = gmail_password.get().strip())
            except:
                mails.wrong_password = tkinter.Label(text = "Incorrect Gmail Password",fg = "red",font = ("Consolas"), background = WINDOWBG2)
                mails.wrong_password.place(x = 200, y = 165)
            else:
                for i in range(2, sheet.max_row + 1):
                    if mails.state.get():
                        mails.msg = EmailMessage()
                        with open(f"{mails.attachment_path}", "rb") as file:
                            data = file.read()
                            attachedfile_name = os.path.basename(mails.attachment_path)
                            mails.msg.add_attachment(data, maintype = "application", subtype = "octet-stream", filename = attachedfile_name)
                        mails.msg["From"] =  from_address.get(1.0, END).strip()
                        mails.msg["Subject"] = mails.subject.get(1.0,END).strip()
                        unique_messages = email_message
                        for field_name, column_number in mailbody_check.items():
                            unique_messages = unique_messages.replace(str(field_name), str(sheet.cell(row = i, column = column_number).value))
                        mails.msg["To"] = sheet.cell(row = i, column = mailbody_check["emailid"]).value
                        mails.msg.attach(MIMEText(unique_messages,'plain'))
                        try:
                            connection.send_message(mails.msg)
                        except:
                            pass
                        del mails.msg["To"]
                    else:
                        unique_messages = email_message
                        for field_name, column_number in mailbody_check.items():
                            unique_messages = unique_messages.replace(str(field_name), str(sheet.cell(row = i, column = column_number).value))
                        try:
                            connection.sendmail(from_addr = from_address.get(1.0, END).strip(), to_addrs = sheet.cell(row = i, column = mailbody_check["emailid"]).value,msg = f"Subject : {mails.subject.get(1.0,END).strip()}\n\n{unique_messages}")
                        except:
                            pass
                mails.completed_label = tkinter.Label(text = "Process Complete",font = ("Consolas"), background = WINDOWBG2)
                mails.completed_label.place(x = 200, y = 165)

    def ConfirmingMails():
        if mails.yesorno.get():
            mails.confirmation.destroy()
            mails.yes_radio.destroy()
            mails.no_radio.destroy()
            mails.field_options.destroy()
            Write_Mails_Final()
        else:
            return
    
    def Write_Mails():
        mails.notfilled = []
        if mails.subject.get(1.0,END).strip() == "" or mails.mail_body.get(1.0, END).strip() == "":
            if mails.subject.get(1.0,END).strip() == "":
                mails.notfilled.append("Subject")
            if mails.mail_body.get(1.0, END).strip() == "":
                mails.notfilled.append("Body of the mail")
        if len(mails.notfilled) != 0:
            mails.confirmation = tkinter.Label(text = "Wish to mail without the following fields? : ",background = WINDOWBG2)
            mails.confirmation.place(x = 50, y = 275)
            mails.fieldsnotfilled = tkinter.StringVar()
            mails.fieldsnotfilled.set(mails.notfilled[0])
            mails.field_options = tkinter.OptionMenu(window, mails.fieldsnotfilled, *mails.notfilled)
            mails.field_options.place(x = 295, y = 272)
            mails.yesorno = tkinter.IntVar()
            mails.yes_radio = tkinter.Radiobutton(text = "Yes",value = 1, variable = mails.yesorno,background=WINDOWBG2, command = ConfirmingMails)
            mails.yes_radio.place(x = 50,y = 300)
            mails.no_radio = tkinter.Radiobutton(text = "No",value = 0, variable = mails.yesorno,background=WINDOWBG2, command = ConfirmingMails)
            mails.no_radio.place(x = 100,y = 300)
        else:
            Write_Mails_Final()
    browsed_file = tkinter.Button(text = "Browse Files : ", borderwidth = 0, command = Browse_File, background = WINDOWBG2)
    browsed_file.place(x = 50, y = 100)
    mails.nextpage_img = Image.open("Images/NextPage.png")
    mails.nextpage_img = ImageTk.PhotoImage(mails.nextpage_img)
    mails.next_button = tkinter.Button(image = mails.nextpage_img, command = mails.NextPage, borderwidth = 0, background = WINDOWBG2)
    mails.next_button.place(x = 650, y = 250)
    file_name = tkinter.Text(height = 1, width = 76, font = ("Consolas",10,"bold"), borderwidth = 0)
    file_name.place(x = 140, y = 100)
    text_label = tkinter.Label(text = "The following fields were identified from The Excel Sheet : ", background = WINDOWBG2).place(x = 50, y = 120)
    displaying_fields = tkinter.Text(height = 5, width = 35, font = ("Consolas",10,"bold"), borderwidth = 0)
    displaying_fields.place(x = 50, y = 150)
    subject_label = tkinter.Label(text = "Subject : ", background = WINDOWBG2).place(x = 50, y = 245)
    mails.subject = tkinter.Text(height = 1, width = 45, font = ("Consolas",12,"bold"), borderwidth = 0)
    mails.subject.place(x = 115, y = 244)
    fromaddress_label = tkinter.Label(text = "From : ", background = WINDOWBG2).place(x = 300, y = 160)
    from_address = tkinter.Text(height = 1, width = 28, font = ("Consolas",10,"bold"), borderwidth = 0)
    from_address.insert(END, f"{on_home.username}")
    from_address.place(x = 350, y = 160)
    gmail_password_label = tkinter.Label(text = "Enter your Gmail Password : ", background = WINDOWBG2)
    gmail_password_label.place(x = 300 ,y = 180)
    gmail_password = tkinter.Entry(width = 20, font = ("Consolas",8,"bold"), borderwidth = 0, show = "*")
    gmail_password.place(x = 470 ,y = 185)
    state_check = tkinter.IntVar()
    show_password = tkinter.Checkbutton(variable = state_check, text = "Show Password", command = DisplayGmailPassword, background = WINDOWBG2)
    show_password.place(x = 580, y = 180)
    toaddress_label = tkinter.Label(text = "To : ", background = WINDOWBG2).place(x = 300, y = 206)
    to_address = tkinter.Text(height = 1, width = 30, font = ("Consolas",10,"bold"), borderwidth = 0)
    to_address.place(x = 350, y = 210)
    enter_body = tkinter.Label(text = "Enter the body of the email : ", background = WINDOWBG2)
    enter_body.place(x = 50, y = 275)
    mails.mail_body = tkinter.Text(height = 7, width = 75, font = ("Consolas",11,"bold"), borderwidth = 0)
    mails.mail_body.place(x = 50, y = 295)
    def RemoveHelp():
        mails.help.destroy()
    def button_clicked():
        mails.n += 1
        if(mails.n % 2 != 0):
            mails.help = tkinter.Text(height = 2, width = 95,background = WINDOWBG3, font = ("Consolas",7,"bold"), borderwidth = 0)
            mails.help.insert(END, "To include variable fields from the excel file, enclose the fields within '<< >>'\nEg : <<Name>>, Would include all the names mentioned in the Names field of the Excel Sheet")
            mails.help.place(x = 68, y = 432)
            window.mainloop()
        elif(mails.n % 2 == 0):
            RemoveHelp()

    mails.helpimg = Image.open("Images/Help.png")
    mails.helpimg = ImageTk.PhotoImage(mails.helpimg)
    helpbutton=tkinter.Button(image = mails.helpimg,command=button_clicked, borderwidth = 0, background = WINDOWBG2)
    helpbutton.place(x = 48, y = 435)
    mails.home_img = Image.open("Images/Home.png")
    mails.home_img = ImageTk.PhotoImage(mails.home_img)
    homebutton = tkinter.Button(image = mails.home_img, command = on_home.MainMenu, background = WINDOWBG2,borderwidth = 0)
    homebutton.pack(side = TOP, anchor = NE)
    loggedin = tkinter.Label(text = f"Logged in as {on_home.username}", background = WINDOWBG2).place(x = 470, y= 480)

def Modify_Excel():
    class ExcelModify:
        def __init__(self):
            self.refbutton = tkinter.Button()
            self.ref_filename = tkinter.Text()
            self.Fields_available = tkinter.Listbox()
            self.name = "../"
            self.n = 0

        def ModifiedExcel(self):
            self.wb2 = openpyxl.load_workbook(modify.name)
            self.sheet2 = self.wb2.active
            clen= self.sheet2.max_column
            rlen = self.sheet2.max_row
            text = ""
            self.sheet2.cell(row = 1, column = clen + 1, value = f"{self.prompting_fieldname.get(1.0, END)}")
            with open(self.referencename) as file:
                content = file.readlines()
            for i in range(2, rlen + 1):
                text = self.sheet2.cell(row = i, column = self.column).value
                for j in range(len(content)):
                    if ((content[j].strip().split('->',2)[0])) == text:
                        d = self.sheet2.cell(row = i, column = clen + 1)
                        d.value = content[j].strip().split('->',2)[1]
            self.wb2.save(self.name)
            modify.finalmsg = tkinter.Label(text = "Process Complete", font = ("Consolas"), background = WINDOWBG2)
            modify.finalmsg.place(x = 275, y = 450)

        def Browse_Reference(self):
            self.referencename = filedialog.askopenfilename(filetypes = (("Notepad Files", "*.txt"),("All files", "*")))
            self.ref_filename = tkinter.Text(height = 1, width = 90, borderwidth = 0, font = ("Consolas",10,"bold"))
            self.ref_filename.insert(END, f"{self.referencename}")
            self.ref_filename.place(x = 200, y = 320)
            self.enter_name = tkinter.Label(text = "Enter the field name, in which you wish to add details : ", background = WINDOWBG2).place(x = 200, y = 350)
            self.prompting_fieldname = tkinter.Text(height = 1, width = 60, font = ("Consolas",10,"bold"), borderwidth = 0)
            self.prompting_fieldname.place(x = 200, y = 375)
            def RemoveHelp():
                modify.help.destroy()
            def button_clicked():
                modify.n += 1
                if(modify.n % 2 != 0):
                    modify.help = tkinter.Text(height = 5, width = 34,background = WINDOWBG3, font = ("Consolas",7,"bold"), borderwidth = 0)
                    modify.help.insert(END, "Ensure that the value being added and the field value with which it is being compared are separated by a '->'(no spaces)")
                    modify.help.place(x = 220, y = 395)
                    window.mainloop()
                elif(modify.n % 2 == 0):
                    RemoveHelp()

            modify.helpimg = Image.open("Images/Help.png")
            modify.helpimg = ImageTk.PhotoImage(modify.helpimg)
            modify.helpbutton=tkinter.Button(image = modify.helpimg,command=button_clicked, borderwidth = 0, background = WINDOWBG2)
            modify.helpbutton.place(x = 200, y = 395) 
            self.submit_fieldname = tkinter.Button(text = "Submit", borderwidth = 0, command = self.ModifiedExcel)
            self.submit_fieldname.place(x = 629, y = 374)

            

        def Field_Selected(self,event):
            self.column = self.Sheet_Parameters[self.Fields_available.get(self.Fields_available.curselection())]
            self.refbutton = tkinter.Button(text = "Choose a Reference File (.txt) format:",background = WINDOWBG2, command = modify.Browse_Reference, borderwidth = 0)
            self.refbutton.place(x = 200, y = 290)

        def Open_Drive(self):
            chrome_driver_path = insert_your_path
            driver = webdriver.Chrome(executable_path = chrome_driver_path)
            driver.get("https://accounts.google.com/signin/v2/identifier?service=writely&sacu=1&rip=1&flowName=GlifWebSignIn&flowEntry=ServiceLogin")
            login = driver.find_element_by_id("identifierId")
            emailid = on_home.username
            login.send_keys(emailid)
            next_button = driver.find_element_by_xpath('//*[@id="identifierNext"]/div/button/span')
            next_button.click()

        def browse_file(self):
            self.Fields_available.destroy()
            setattr(modify,'name',filedialog.askopenfilename(filetypes = (("Excel Files", "*.xlsx"),("All files", "*"))))
            self.wb = openpyxl.load_workbook(modify.name)
            if file_name.get(1.0,END) != "":
                file_name.delete(1.0,END)
            file_name.insert(END,f"{modify.name}")
            self.sheet = self.wb.active 
            self.Sheet_Parameters = {}
            self.column = self.sheet.max_column
            self.max_column = self.sheet.max_column
            self.row = self.sheet.max_row
            for i in range(1, self.column + 1):
                self.Sheet_Parameters[self.sheet.cell(row = 1, column = i).value] = i
            self.wb.close()
            label = tkinter.Label(text = "The following fields have been identified.\nChoose which field you wish to compare with : ", background = WINDOWBG2).place(x = 40, y = 242)
            self.Fields_available = tkinter.Listbox(height = len(self.Sheet_Parameters), borderwidth=0, font = ("Consolas",10,"bold"))
            for key, value in self.Sheet_Parameters.items():
                self.Fields_available.insert(value, key)
                self.Fields_available.bind("<<ListboxSelect>>",modify.Field_Selected)
                self.Fields_available.place(x = 40, y = 290)
    
    modify = ExcelModify()
    Remove()
    modify.browse_img = Image.open("Images/Browse.png").resize((150,150))
    modify.browse_img = ImageTk.PhotoImage(modify.browse_img)
    modify.browse_Button = tkinter.Button(image = modify.browse_img, command = modify.browse_file, borderwidth = 0, background = WINDOWBG2)
    modify.browse_Button.place(x = 175, y = 40)
    browselabel = tkinter.Label(text = "Browse Excel Files", background = WINDOWBG2).place(x = 200, y = 175)
    modify.drive_img = Image.open("Images/GDrive.png").resize((150,150))
    modify.drive_img = ImageTk.PhotoImage(modify.drive_img)
    modify.download_Button = tkinter.Button(image = modify.drive_img, command = modify.Open_Drive,borderwidth = 0, background = WINDOWBG2)
    modify.download_Button.place(x = 350, y = 40)
    drivelabel = tkinter.Label(text = "Download From Drive", background = WINDOWBG2).place(x = 360, y = 175)
    openfile_label = tkinter.Label(text = "File Opened : ", background = WINDOWBG2).place(x = 125, y = 202)
    file_name = tkinter.Text(height = 1, width = 80, font = ("Consolas",8,"bold"), borderwidth = 0)
    file_name.place(x = 210, y = 205)
    modify.home_img = Image.open("Images/Home.png")
    modify.home_img = ImageTk.PhotoImage(modify.home_img)
    homebutton = tkinter.Button(image = modify.home_img, command = on_home.MainMenu, background = WINDOWBG2, borderwidth = 0)
    homebutton.pack(side = TOP, anchor = NE)
    loggedin = tkinter.Label(text = f"Logged in as {on_home.username}", background = WINDOWBG2).pack(side = BOTTOM, anchor = SE)

class Home:

    def __init__(self):
        self.next_img = Image.open("Images/NextPageColored.png")
        self.next_img = ImageTk.PhotoImage(self.next_img)

    def DisplayPassword(self):
        if self.state_check.get() == 1:
            self.passwordwidget.config(show = "")
        else:
            self.passwordwidget.config(show = "*")


    def HomePage(self):
        on_home.homebutton.destroy()
        self.signin_img = Image.open("Images/SignUp.png")
        self.signin_img = ImageTk.PhotoImage(self.signin_img)
        self.signin_button = tkinter.Button(image = self.signin_img, borderwidth = 0, command = on_home.Clicked_Signin)
        self.signin_button.place(x = 60, y = 100)

        self.app_img = Image.open("Images/App name.png")
        self.app_img = ImageTk.PhotoImage(self.app_img)
        self.app = tkinter.Label(image = self.app_img, borderwidth = 0)
        self.app.place(x = 262, y = 100)

        self.login_img = Image.open("Images/Login.png")
        self.login_img = ImageTk.PhotoImage(self.login_img)
        self.login_button = tkinter.Button(image = self.login_img, borderwidth = 0, command = on_home.Clicked_Login)
        self.login_button.place(x = 462, y = 100)


    def MainMenu(self):
        Remove()
        on_home.homebutton.destroy()
        self.sort_buttonimg = Image.open("Images/Excel_Icon.png").resize((125, 125))
        self.sort_buttonimg = ImageTk.PhotoImage(self.sort_buttonimg)
        self.sort_button = tkinter.Button(image = self.sort_buttonimg, borderwidth = 0, command = Form_Sorting,background = WINDOWBG2).place(x = 90, y = 41)
        self.sort_dimg = Image.open("Images/Sorting Description.png").resize((250,125))
        self.sort_dimg = ImageTk.PhotoImage(self.sort_dimg)
        self.sort_description = tkinter.Label(image = self.sort_dimg, borderwidth = 0).place(x = 220, y = 43)

        self.mail_buttonimg = Image.open("Images/Email.png").resize((125, 125))
        self.mail_buttonimg = ImageTk.PhotoImage(self.mail_buttonimg)
        self.mail_button = tkinter.Button(image = self.mail_buttonimg, borderwidth = 0, command = Send_Mails,background = WINDOWBG2).place(x = 90, y = 350)
        self.mailimg = Image.open("Images/Mailing Description.png").resize((250,125))
        self.mailimg = ImageTk.PhotoImage(self.mailimg)
        self.mail_description = tkinter.Label(image = self.mailimg, borderwidth = 0).place(x = 208, y = 350)

        self.modify_buttonimg = Image.open("Images/ModifyExcelicon.png").resize((125, 125))
        self.modify_buttonimg = ImageTk.PhotoImage(self.modify_buttonimg)
        self.modify_button = tkinter.Button(image = self.modify_buttonimg, borderwidth = 0, command = Modify_Excel,background = WINDOWBG2).place(x = 90, y = 196)
        self.mimg = Image.open("Images/Modifier Description.png").resize((250,125))
        self.mimg = ImageTk.PhotoImage(self.mimg)
        self.modify_description = tkinter.Label(image = self.mimg, borderwidth = 0).place(x = 219, y = 200)

        self.loggedin = tkinter.Label(text = f"Logged in as {self.username}", background = WINDOWBG2).pack(side = BOTTOM, anchor = SE)


    def CreatingAccount(self):
        if self.passwordwidget.get() != self.confirm_passwordwidget.get():
            self.different_passwords = tkinter.Label(text = "Please enter the password entered earlier : ", fg = "red", background = WINDOWBG2).place(x = 310, y = 340)
        else:
            f=open("logindetails.json","r+")
            contents = f.read()
            js = json.loads(contents)
            js.append({'username': self.username, 'password': self.passwordwidget.get().strip()})
            f.seek(0)       
            f.write(json.dumps(js, indent=2))
            f.truncate()
            f.close()
            self.MainMenu()
    
    def CheckingForAccount(self):
        f = open("logindetails.json","r+")
        contents = f.read()
        js = json.loads(contents)
        self.email = True
        for i in range(len(js)):
            if js[i]['username'] == self.username:
                if js[i]['password'] == self.password:
                    self.email = True
                    self.MainMenu()
                else:
                    self.incorrectpassword = tkinter.Label(text = "Incorrect Password", borderwidth = 0, background = WINDOWBG2, fg = "red")
                    self.incorrectpassword.place(x = 325, y = 275)
                    break
            else:
                self.email = False  
        if self.email is False:
            self.incorrectemail = tkinter.Label(text = "Email does not exist, try to Sign Up",  borderwidth = 0, background = WINDOWBG2, fg = "red").place(x = 110, y = 230)
            self.usernamewidget.config(fg = "red")
        f.close()

    def Password_Confirmation(self):
        if self.password != self.verificationwidget.get(1.0, END).strip():
            self.verificationwidget.config(font = ("red"))
            self.incorrectid = tkinter.Label(text = "Incorrect Authentication id", fg = "red", background = WINDOWBG2).place(x = 310, y = 330)
        else:
            self.userdetails_img = Image.open("Images/SignupPage 2.png")
            self.userdetails_img = ImageTk.PhotoImage(self.userdetails_img)
            self.user = tkinter.Label(image = self.userdetails_img, borderwidth = 0)
            self.user.place(x = 262, y = 100)
            self.passwordwidget = tkinter.Entry(width = 25, font = ("Consolas",14,"bold"), borderwidth = 0, show = "*", background = "#B2FFFF", bd = 3)
            self.passwordwidget.place(x = 310, y  = 200)
            self.state_check = tkinter.IntVar()
            self.show_password = tkinter.Checkbutton(variable = self.state_check, text = "Show Password :", command = self.DisplayPassword, background = "#B2FFFF")
            self.show_password.place(x = 310, y = 230)
            self.confirm_passwordwidget = tkinter.Entry(width = 25, font = ("Consolas",14,"bold"), borderwidth = 0, show = "*", background = "#B2FFFF", bd = 4)
            self.confirm_passwordwidget.place(x = 310, y  = 300)           
            self.next3 = tkinter.Button(image = self.next_img, borderwidth = 0, command = self.CreatingAccount, background = "#B2FFFF")
            self.next3.place(x = 590, y = 275)

    def Submit_NewUser(self):
        def check_mail(input_username):
            if(re.search(regex,input_username)):   
                return True   
            else:   
                return False
        self.isvalid = check_mail(self.usernamewidget.get(1.0, END).strip())
        if self.isvalid:
            self.homebutton.destroy()
            app.destroy(); login_button.destroy()
            self.username = self.usernamewidget.get(1.0, END).strip()
            self.lower=string.ascii_lowercase
            self.upper=string.ascii_uppercase
            self.num=string.digits
            all = self.lower + self.upper + self.num   
            temp=random.sample(all,5)
            self.password="".join(temp)
            self.userdetails_img = Image.open("Images/SignupPage 3.png")
            self.userdetails_img = ImageTk.PhotoImage(self.userdetails_img)
            self.user = tkinter.Label(image = self.userdetails_img, borderwidth = 0)
            self.user.place(x = 262, y = 100)
            self.verificationwidget = tkinter.Text(height = 1, width = 25, font = ("Consolas",14,"bold"), borderwidth = 0, background = "#B2FFFF", bd = 4)
            self.verificationwidget.place(x = 310, y  = 270)
            self.next2 = tkinter.Button(image = self.next_img, borderwidth = 0, command = self.Password_Confirmation, background = "#B2FFFF")
            self.next2.place(x = 590, y = 235)
            with smtplib.SMTP("smtp.gmail.com") as server:
                server.starttls()
                server.login('testproject194@gmail.com', 'testproject!@#$')
                server.sendmail(from_addr = 'testproject194@gmail.com', to_addrs = self.username, msg = f"Subject : Authentication Code\n\nAuthentication Code is {self.password}")
        else:
            self.notvalid = tkinter.Label(text = "Not a Valid Email id", fg = "red", background = "#B2FFFF").place(x = 310, y = 260)


    def Submit_Login(self):
        self.username = self.usernamewidget.get(1.0, END).strip()
        self.password = self.passwordwidget.get().strip()
        self.CheckingForAccount()
    
    def Clicked_Signin(self):
        self.homebutton.destroy()
        app.destroy(); login_button.destroy()
        self.userdetails_img = Image.open("Images/SignupPage 1.png")
        self.userdetails_img = ImageTk.PhotoImage(self.userdetails_img)
        self.user = tkinter.Label(image = self.userdetails_img, borderwidth = 0)
        self.user.place(x = 262, y = 100)
        self.usernamewidget = tkinter.Text(height = 1, width = 30, font = ("Consolas",14,"bold"), borderwidth = 0, background = "#B2FFFF", bd = 4)
        self.usernamewidget.place(x = 310, y  = 200)
        self.next1 = tkinter.Button(image = self.next_img, borderwidth = 0, command = self.Submit_NewUser, background = "#B2FFFF")
        self.next1.place(x = 590, y = 235)
        self.home_img = Image.open("Images/Home.png")
        self.home_img = ImageTk.PhotoImage(self.home_img)
        self.homebutton = tkinter.Button(image = self.home_img, command = on_home.HomePage, background = WINDOWBG2,borderwidth = 0)
        self.homebutton.pack(side = TOP, anchor = NE)


    def Clicked_Login(self):
        self.homebutton.destroy()
        app.destroy(); signin_button.destroy()
        self.userdetails_img = Image.open("Images/UserDetails.png")
        self.userdetails_img = ImageTk.PhotoImage(self.userdetails_img)
        self.user = tkinter.Label(image = self.userdetails_img, borderwidth = 0)
        self.user.place(x = 60, y = 100)
        self.usernamewidget = tkinter.Text(height = 1, width = 30, font = ("Consolas",14,"bold"), borderwidth = 0,bd = 4, background = "#B2FFFF")
        self.usernamewidget.place(x = 110, y  = 200)
        self.passwordwidget = tkinter.Entry(width = 30, font = ("Consolas",14,"bold"), borderwidth = 0, show = "*",background = "#B2FFFF",bd = 4)
        self.passwordwidget.place(x = 110, y  = 295)
        self.state_check = tkinter.IntVar()
        self.show_password = tkinter.Checkbutton(variable = self.state_check, text = "Show Password :", command = self.DisplayPassword,background = "#B2FFFF")
        self.show_password.place(x = 110, y = 340)
        self.next = tkinter.Button(image = self.next_img, borderwidth = 0, command = self.Submit_Login, background = "#B2FFFF")
        self.next.place(x = 389, y = 235)
        self.home_img = Image.open("Images/Home.png")
        self.home_img = ImageTk.PhotoImage(self.home_img)
        self.homebutton = tkinter.Button(image = self.home_img, command = on_home.HomePage, background = WINDOWBG2, borderwidth = 0)
        self.homebutton.pack(side = TOP, anchor = NE)

on_home = Home()


signin_img = Image.open("Images/SignUp.png")
signin_img = ImageTk.PhotoImage(signin_img)
signin_button = tkinter.Button(image = signin_img, borderwidth = 0, command = on_home.Clicked_Signin)
signin_button.place(x = 60, y = 100)

app_img = Image.open("Images/App name.png")
app_img = ImageTk.PhotoImage(app_img)
app = tkinter.Label(image = app_img, borderwidth = 0)
app.place(x = 262, y = 100)

login_img = Image.open("Images/Login.png")
login_img = ImageTk.PhotoImage(login_img)
login_button = tkinter.Button(image = login_img, borderwidth = 0, command = on_home.Clicked_Login)
login_button.place(x = 462, y = 100)

home_img = Image.open("Images/Home.png")
home_img = ImageTk.PhotoImage(home_img)
on_home.homebutton = tkinter.Button(image = home_img, command = on_home.HomePage)

window.mainloop()
