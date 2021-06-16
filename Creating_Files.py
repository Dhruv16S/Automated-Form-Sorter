import openpyxl
import tkinter
from tkinter import filedialog
class TkinterReturns:
    def __init__(self):
        self.column = 0
        self.name = "."
    def Field_Selected(self,event):
        self.column = Sheet_Parameters[Fields_available.get(Fields_available.curselection())]
        Creating_Files(self.column)
    def browse_file(self):
        self.name = filedialog.askopenfilename(filetypes = (("All files", "*"), ("Template files", "*.type")))
        window.destroy()

fields = TkinterReturns()

window = tkinter.Tk()
window.minsize(height = 500, width = 500)
browse_Button = tkinter.Button(master = window, text = 'Browse', width = 6, command=fields.browse_file, borderwidth = 0)
browse_Button.pack(side=tkinter.LEFT, padx = 2, pady=2)
window.mainloop()


wb = openpyxl.load_workbook(fields.name)   
window = tkinter.Tk()
window.minsize(height = 500, width = 500)


sheet = wb.active 
Sheet_Parameters = {}
column = sheet.max_column
row = sheet.max_row

for i in range(1, column + 1):
    Sheet_Parameters[sheet.cell(row = 1, column = i).value] = i


label = tkinter.Label(window, text = "The following fields have been identified, how would you like to classify the form : ")
label.pack()
Fields_available = tkinter.Listbox(height = len(Sheet_Parameters), borderwidth=0)
for key, value in Sheet_Parameters.items():
    Fields_available.insert(value, key)
Fields_available.bind("<<ListboxSelect>>",fields.Field_Selected)
Fields_available.pack()

def Creating_Files(operating_column):
    for i in range(2, row + 1):    
        file_name = sheet.cell(i, operating_column).value.split(", ")      #Do I want to split on the basis of number also....
        for field_types in file_name:                                      
            try:
                new_wb = openpyxl.load_workbook(f"{field_types}.xlsx")
            except FileNotFoundError:
                new_wb = openpyxl.Workbook()
            new_sheet = new_wb.active
            new_rows = new_sheet.max_row + 1
            for j in range(1, column + 1):
                new_sheet.cell(new_rows,j,value = sheet.cell(i,j).value)    
            new_wb.save(f"{field_types}.xlsx")
window.mainloop()
