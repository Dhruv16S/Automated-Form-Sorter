import openpyxl
user = input("Enter a name")
Sheet_Fields = {}
try:
    wb = openpyxl.load_workbook('Book1.xlsx')
except:
    print("the file does not exist")
else:
    sheet = wb.active
    with open("link.txt") as file:
        content = file.readlines()
    clen= sheet.max_column
    rlen = sheet.max_row
    text = ""
    for i in range(1, clen + 1):
        Sheet_Fields[sheet.cell(row = 1, column = i).value] = i
    print(Sheet_Fields)
    compared_field = input("Enter a field :")
    if compared_field not in Sheet_Fields.keys():
        print("Incorrect")
    else:
        sheet.cell(row = 1, column = clen + 1, value = f"{user}")
        for i in range(2, rlen + 1):
            text = sheet.cell(row = i, column = Sheet_Fields[compared_field]).value
            for j in range(len(content)):
                if ((content[j].strip().split('-')[0])) == text:
                    d = sheet.cell(row = i, column = clen + 1)
                    d.value = content[j].strip().split('-')[1]

        wb.save('Book1.xlsx')