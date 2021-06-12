
# Get access to drive link
# Read number of rows, then access heading of each column. According to users interest, choose what to select

import openpyxl
wb = openpyxl.load_workbook("Sample.xlsx")   # Enter file name here
sheet = wb['Sheet1']   #Enter sheet name here
Sheet_Parameters = {}
column = sheet.max_column
row = sheet.max_row
for i in range(1, column + 1):
    Sheet_Parameters[sheet.cell(row = 1, column = i).value] = i
print(Sheet_Parameters)
choice = input("The following fields were obtained how would you like to classify the form : ?")
operating_column = Sheet_Parameters[choice]
for i in range(2, row + 1):    # Don't begin from 1 as name of the field also comes, while dividing for similar options check if string
    # print(sheet.cell(i, operating_column).value)
    file_name = sheet.cell(i, operating_column).value.split(",")
    for field_types in file_name:
        try:
            new_wb = openpyxl.load_workbook(f"{field_types}.xlsx")
        except FileNotFoundError:
            new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_rows = new_sheet.max_row + 1
        for j in range(1, column + 1):
            new_sheet.cell(new_rows,j,value = sheet.cell(i,j).value)    
        new_wb.save(f"{field_types}.xlsx")